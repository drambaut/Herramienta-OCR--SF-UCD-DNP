"""
Herramienta OCR-SF-UCD V13.0 — DNP
====================================================================
Procesamiento integral de expedientes de pago a contratistas.

Pipeline:
  1. Recibe un ZIP con PDFs mezclados (RAS, PILA, Facturas, AFC,
     Dependientes, Formato 5, Certificación Bancaria, Registro Civil).
  2. OCR con Azure Document Intelligence + clasificación / extracción
     con Azure OpenAI (GPT-4o visión).
  3. Vinculación relacional Late-Binding: agrupa por cédula → mes_cobro.
  4. Llena la plantilla Excel RT-Cto1.xlsx con los datos del expediente.
  5. Empaqueta todos los Excels en un ZIP descargable
     (RT-Cto-{cedula}.xlsx).

Subdirección Financiera · Unidad de Costos y Dependencias · DNP
"""

import os
import io
import json
import base64
import time
import zipfile
import tempfile
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

import pandas as pd
import fitz                                    # PyMuPDF
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from openai import AzureOpenAI
from dotenv import load_dotenv

load_dotenv()

# ════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ════════════════════════════════════════════════════════════════════
ENDPOINT                 = os.getenv("AZURE_FORM_RECOGNIZER_ENDPOINT")
API_KEY                  = os.getenv("AZURE_FORM_RECOGNIZER_KEY")
AZURE_OPENAI_ENDPOINT    = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_KEY     = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-05-01-preview")
AZURE_OPENAI_DEPLOYMENT  = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")

# Plantilla Excel base (debe estar comprometida en el repo)
RUTA_PLANTILLA = Path(__file__).parent / "template" / "RT-Cto1.xlsx"

# Parámetros tributarios (referencia; ya viven dentro de la plantilla)
UVT_2026   = 52379
SMLMV_2026 = 1750905.0

# ════════════════════════════════════════════════════════════════════
# ESTILOS GLOBALES (heredados de V1)
# ════════════════════════════════════════════════════════════════════
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
    background-color: #070c18;
    color: #c9d4e8;
}
.stApp {
    background: linear-gradient(160deg, #070c18 0%, #0d1525 50%, #070c18 100%);
    min-height: 100vh;
}
.ocr-header {
    padding: 2.5rem 0 1.5rem 0;
    border-bottom: 1px solid #1a2744;
    margin-bottom: 2rem;
}
.ocr-header h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem; font-weight: 600;
    color: #ffffff; letter-spacing: -0.02em; margin: 0;
}
.ocr-header .subtitle {
    font-size: 0.82rem; color: #4a6090;
    letter-spacing: 0.12em; text-transform: uppercase; margin-top: 0.4rem;
}
.ocr-badge {
    display: inline-block;
    background: #0f2044; border: 1px solid #1e3a6e; color: #3b82f6;
    font-family: 'IBM Plex Mono', monospace; font-size: 0.68rem;
    padding: 0.2rem 0.7rem; border-radius: 2px;
    letter-spacing: 0.1em; margin-bottom: 0.8rem;
}
.kpi-card {
    background: linear-gradient(135deg, #0d1830 0%, #111e38 100%);
    border: 1px solid #1a2d55; border-radius: 10px;
    padding: 1.6rem 1.4rem; position: relative; overflow: hidden;
}
.kpi-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #1e3a6e, #3b82f6, #1e3a6e);
}
.kpi-value {
    font-family: 'IBM Plex Mono', monospace; font-size: 2.4rem;
    font-weight: 600; color: #3b82f6; line-height: 1; margin-bottom: 0.5rem;
}
.kpi-label {
    font-size: 0.68rem; letter-spacing: 0.14em; text-transform: uppercase;
    color: #4a6090; font-weight: 600;
}
.kpi-sub {
    font-size: 0.75rem; color: #2a4070; margin-top: 0.3rem;
    font-family: 'IBM Plex Mono', monospace;
}
.progress-wrapper {
    background: #0a1022; border: 1px solid #1a2744; border-radius: 8px;
    padding: 1.8rem 2rem; margin: 1.5rem 0;
}
.progress-title {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.78rem;
    color: #3b82f6; letter-spacing: 0.1em; text-transform: uppercase;
    margin-bottom: 1rem;
}
.progress-track {
    background: #0d1525; border-radius: 4px; height: 6px;
    overflow: hidden; position: relative;
}
.progress-fill {
    height: 100%; border-radius: 4px;
    background: linear-gradient(90deg, #1d4ed8, #3b82f6, #60a5fa);
    background-size: 200% 100%;
    animation: shimmer 1.4s ease infinite;
    transition: width 0.6s ease;
}
@keyframes shimmer {
    0%   { background-position: 200% center; }
    100% { background-position: -200% center; }
}
.progress-pulse {
    display: flex; align-items: center; gap: 0.6rem; margin-top: 0.9rem;
}
.pulse-dot {
    width: 7px; height: 7px; background: #3b82f6; border-radius: 50%;
    animation: pulse 1.2s ease-in-out infinite; flex-shrink: 0;
}
@keyframes pulse {
    0%, 100% { opacity: 1; transform: scale(1); }
    50%      { opacity: 0.3; transform: scale(0.6); }
}
.pulse-dot:nth-child(2) { animation-delay: 0.2s; }
.pulse-dot:nth-child(3) { animation-delay: 0.4s; }
.progress-msg {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; color: #4a6090;
}
.progress-counter {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.72rem;
    color: #2a4070; margin-top: 0.5rem; text-align: right;
}
.section-title {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.72rem;
    letter-spacing: 0.14em; text-transform: uppercase; color: #2a4880;
    border-bottom: 1px solid #111e38; padding-bottom: 0.6rem;
    margin: 2rem 0 1.2rem 0;
}
.stDownloadButton > button {
    background: linear-gradient(135deg, #1d4ed8, #2563eb) !important;
    color: white !important; border: none !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.78rem !important; letter-spacing: 0.08em !important;
    padding: 0.6rem 1.4rem !important; transition: all 0.2s !important;
    box-shadow: 0 0 20px rgba(59,130,246,0.15) !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #2563eb, #3b82f6) !important;
    box-shadow: 0 0 30px rgba(59,130,246,0.3) !important;
    transform: translateY(-1px) !important;
}
[data-testid="stFileUploader"] {
    background: #0a1022; border: 1px dashed #1e3a6e;
    border-radius: 8px; padding: 1rem;
}
[data-testid="stFileUploader"]:hover { border-color: #3b82f6; }
.stAlert {
    background: #0a1022 !important; border-color: #1a2744 !important;
    color: #c9d4e8 !important;
}
.stDataFrame { border: 1px solid #1a2744; border-radius: 6px; overflow: hidden; }
::-webkit-scrollbar { width: 4px; }
::-webkit-scrollbar-track { background: #070c18; }
::-webkit-scrollbar-thumb { background: #1a2744; border-radius: 2px; }
.success-banner {
    background: linear-gradient(135deg, #052e16, #064e3b);
    border: 1px solid #065f46; border-radius: 8px;
    padding: 1rem 1.4rem; color: #6ee7b7;
    font-family: 'IBM Plex Mono', monospace; font-size: 0.8rem;
    margin: 1rem 0;
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1rem; max-width: 1100px; }
</style>
"""

# ════════════════════════════════════════════════════════════════════
# CLIENTES AZURE
# ════════════════════════════════════════════════════════════════════
@st.cache_resource
def get_clients():
    cliente_di  = DocumentIntelligenceClient(ENDPOINT, AzureKeyCredential(API_KEY))
    cliente_oai = AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        api_key=AZURE_OPENAI_API_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
    )
    return cliente_di, cliente_oai


@st.cache_data(show_spinner=False)
def cargar_plantilla_bytes() -> bytes:
    """Lee la plantilla Excel desde disco una sola vez por sesión."""
    if not RUTA_PLANTILLA.exists():
        return b""
    return RUTA_PLANTILLA.read_bytes()


# ════════════════════════════════════════════════════════════════════
# UTILIDADES (parseo de fechas y números)
# ════════════════════════════════════════════════════════════════════
def parse_fecha(valor):
    """Convierte un string YYYY-MM-DD (o variantes) a datetime; tolerante."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor.strip()[:10], fmt)
            except ValueError:
                continue
    return None


def parse_numero(valor):
    """Convierte cualquier valor a float; tolerante a strings con $ , ."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        limpio = valor.replace("$", "").replace("COP", "").replace(" ", "")
        limpio = limpio.replace(".", "").replace(",", ".")
        try:
            return float(limpio)
        except ValueError:
            return None
    return None


def mes_desde_fecha(valor):
    f = parse_fecha(valor)
    return f.month if f else None


def add_years_safe(d: datetime, years: int) -> datetime:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d.replace(year=d.year + years, day=28)


# ════════════════════════════════════════════════════════════════════
# OCR — AZURE DOCUMENT INTELLIGENCE
# ════════════════════════════════════════════════════════════════════
def pdf_primera_pagina_a_base64(ruta_pdf: str, dpi: int = 150) -> str:
    doc = fitz.open(ruta_pdf)
    pag = doc[0]
    pix = pag.get_pixmap(dpi=dpi)
    data = pix.tobytes("png")
    doc.close()
    return base64.standard_b64encode(data).decode("utf-8")


def extraer_texto_con_azure(ruta_pdf: str, cliente_di) -> str:
    with open(ruta_pdf, "rb") as f:
        tarea = cliente_di.begin_analyze_document(
            "prebuilt-layout", body=f, content_type="application/octet-stream"
        )
        resultado = tarea.result()

    bloques = []
    for pagina in resultado.pages:
        for linea in pagina.lines:
            bloques.append(linea.content)

    if resultado.tables:
        bloques.append("\n--- TABLAS DETECTADAS ---")
        for tabla in resultado.tables:
            for celda in tabla.cells:
                bloques.append(
                    f"[Fila {celda.row_index}, Col {celda.column_index}]: {celda.content}"
                )
    return "\n".join(bloques)


# ════════════════════════════════════════════════════════════════════
# IA — CLASIFICACIÓN Y EXTRACCIÓN UNIVERSAL
# ════════════════════════════════════════════════════════════════════
SYSTEM_PROMPT = """
Eres el Auditor Financiero Principal del DNP. Procesas documentos de pago a contratistas:
RAS, PLANILLA (PILA), FACTURA, FORMATO_5, AFC, DEPENDIENTES.

Devuelve ÚNICAMENTE un objeto JSON válido, sin texto adicional, sin markdown.

REGLAS:
1. VALORES MONETARIOS: solo números puros (float). Sin $, sin "COP", sin separadores.
2. MESES: convierte cualquier mes (Enero, Feb, 03) a NÚMERO ENTERO del 1 al 12.
3. OPERADOR PILA: ignora bancos (Bancolombia, etc.). Solo: SOI, MiPlanilla,
   Aportes en Linea, ARUS, Asopago, Simple, Enlace Operativo.
4. FECHAS: formato YYYY-MM-DD siempre.
5. IDENTIFICACIONES: solo dígitos, sin puntos, sin guiones.
6. Si un dato no aplica al tipo de documento, devuelve null.
"""

USER_PROMPT_TEMPLATE = """
Clasifica el documento en UNO de estos tipos:
  - RAS: Registro/Resumen de Aprobación de Solicitud de pago.
  - PLANILLA: Planilla PILA seguridad social (SOI, MiPlanilla, ARUS, etc.).
  - FACTURA: Factura electrónica de venta DIAN.
  - FORMATO_5: Formato N°5 de cuenta de cobro.
  - AFC: Solicitud "Aportes a Cuentas AFC" / "Ahorro Fomento Construcción".
  - DEPENDIENTES: Solicitud "Deducción por Dependientes" (Art. 387 ET).
  - OTRO: cualquier otro (certificación bancaria, registro civil anexo, etc.).

Extrae los siguientes datos. Si un campo NO aplica, devuelve null.

{{
  "tipo_documento_detectado": "RAS|PLANILLA|FACTURA|FORMATO_5|AFC|DEPENDIENTES|OTRO",

  "identificacion_contratista": "Cédula o NIT (solo dígitos)",
  "nombre_contratista": "Nombre completo o Razón Social del contratista",
  "numero_contrato": "Ej: DNP-206-2026",

  "mes_cobro_honorarios": <Mes 1-12 de los honorarios cobrados>,
  "anio_cobro_honorarios": <Año del cobro>,

  "ras_pago_actual": <Número de este pago, ej: 2 de 9 -> 2>,
  "ras_pago_total": <Total de pagos del contrato, ej: 2 de 9 -> 9>,
  "ras_periodo_inicio": "YYYY-MM-DD",
  "ras_periodo_fin": "YYYY-MM-DD",
  "ras_nombres_supervisor": "Nombre supervisor",
  "ras_valor_dnp": <Monto pagado por DNP>,
  "ras_valor_sgr": <Monto pagado por SGR>,
  "ras_registro_contable_no": "Número de registro contable",
  "ras_fecha_registro": "YYYY-MM-DD",

  "factura_cufe": "CUFE",
  "factura_numero": "Número factura",
  "factura_fecha_emision": "YYYY-MM-DD",
  "factura_valor_sin_iva": <Subtotal sin IVA>,
  "factura_iva": <IVA>,
  "factura_total": <Total con IVA>,

  "pila_fecha_pago": "YYYY-MM-DD",
  "pila_operador": "Operador PILA (sin bancos)",
  "pila_periodo_pagado": <Mes 1-12 que cubre la cotización>,
  "pila_ibc": <IBC>,
  "pila_salud": <Aporte salud>,
  "pila_pension": <Aporte pensión>,
  "pila_arl": <Aporte ARL>,

  "formato5_mes_planilla": <Mes 1-12 de la planilla referida en el formato 5>,
  "formato5_deduccion_renta": "SI|NO",

  "afc_monto": <Monto AFC a deducir>,
  "afc_fecha_solicitud": "YYYY-MM-DD",
  "afc_mes_aplicable": <Mes 1-12 al que aplica la deducción AFC>,
  "afc_banco": "Banco donde está la cuenta AFC",
  "afc_numero_cuenta": "Número de cuenta AFC",

  "dep_nombre": "Nombre completo del dependiente",
  "dep_tipo_doc": "CC|TI|RC|Otro",
  "dep_numero_doc": "Documento del dependiente",
  "dep_fecha_nacimiento": "YYYY-MM-DD",
  "dep_parentesco": "Hijo|Padre|Cónyuge|Hermano|Otro",
  "dep_calidad_dependencia": "Menor de edad|Estudiante|Discapacidad|Sin ingresos|Otro",
  "dep_fecha_solicitud": "YYYY-MM-DD"
}}

TEXTO DEL DOCUMENTO:
{texto}
"""

def procesar_con_ia(texto: str, imagen_b64: str, cliente_oai) -> dict:
    try:
        respuesta = cliente_oai.chat.completions.create(
            model=AZURE_OPENAI_DEPLOYMENT,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": [
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{imagen_b64}",
                                   "detail": "high"}},
                    {"type": "text", "text": USER_PROMPT_TEMPLATE.format(texto=texto[:7000])},
                ]},
            ],
            max_tokens=1500,
            temperature=0.0,
            response_format={"type": "json_object"},
        )
        raw = respuesta.choices[0].message.content.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        return json.loads(raw)
    except Exception as e:
        return {"_error_json": f"Fallo IA: {str(e)[:200]}"}


# ════════════════════════════════════════════════════════════════════
# VINCULACIÓN RELACIONAL (LATE BINDING)
# ════════════════════════════════════════════════════════════════════
def _nuevo_expediente():
    return {
        "identificacion"         : None,
        "nombre"                 : None,
        "numero_contrato"        : None,
        "fecha_inicio_ejecucion" : None,
        "fecha_fin_ejecucion"    : None,
        "valor_mensual"          : None,
        "valor_total"            : None,
        "numero_pagos"           : None,
        "es_pensionado"          : None,
        "responsable_iva"        : None,
        "regimen_simple"         : None,
        "banco"                  : None,
        "numero_cuenta"          : None,
        "supervisor"             : None,
        "dependencia"            : None,
        "pagos_por_mes"          : {},                   # mes(int) → dict
        "afc_por_mes"            : {},                  # mes(int) → monto(float)
        "pilas_por_mes"          : {},                  # mes(int) → dict
        "dependientes"           : [],
        "documentos_detectados"  : [],
        "alertas"                : [],
    }


def construir_expedientes(documentos_procesados: list) -> dict:
    """Agrupa por cédula. Late-binding para PILAs huérfanas."""
    expedientes = defaultdict(_nuevo_expediente)
    pilas_huerfanas = []
    formato5_index  = {}   # (cedula, mes_cobro) → mes_planilla

    # ─── FASE 1: documentos NO-PILA ────────────────────────────────
    for doc in documentos_procesados:
        if "_error_json" in doc:
            continue

        cedula = str(doc.get("identificacion_contratista") or "SIN_CEDULA").strip()
        tipo   = (doc.get("_tipo") or "").upper()
        archivo = doc.get("_archivo", "")

        if "PILA" in tipo or "PLANILLA" in tipo:
            pilas_huerfanas.append(doc)
            continue

        exp = expedientes[cedula]
        exp["identificacion"] = cedula
        if not exp["nombre"] and doc.get("nombre_contratista"):
            exp["nombre"] = doc["nombre_contratista"]
        if not exp["numero_contrato"] and doc.get("numero_contrato"):
            exp["numero_contrato"] = doc["numero_contrato"]
        exp["documentos_detectados"].append({"tipo": tipo or "OTRO", "archivo": archivo})

        # helper para acceder/crear entrada mensual de forma segura
        def pago_mes(m):
            k = int(m)
            if k not in exp["pagos_por_mes"]:
                exp["pagos_por_mes"][k] = {}
            return exp["pagos_por_mes"][k]

        # ── RAS ────────────────────────────────────────────
        if "RAS" in tipo:
            mes = doc.get("mes_cobro_honorarios")
            if mes:
                try:
                    mes_int = int(mes)
                    if 1 <= mes_int <= 12:
                        p = pago_mes(mes_int)
                        if doc.get("ras_valor_dnp")          is not None: p["valor_dnp"]           = parse_numero(doc["ras_valor_dnp"])
                        if doc.get("ras_valor_sgr")          is not None: p["valor_sgr"]           = parse_numero(doc["ras_valor_sgr"])
                        if doc.get("ras_registro_contable_no"):            p["registro_contable_no"] = doc["ras_registro_contable_no"]
                        if doc.get("ras_fecha_registro"):                  p["fecha_registro"]       = doc["ras_fecha_registro"]
                        if doc.get("ras_nombres_supervisor"):              p["supervisor"]            = doc["ras_nombres_supervisor"]
                        if doc.get("ras_pago_actual")         is not None: p["pago_actual"]           = doc["ras_pago_actual"]
                        # El valor del RAS también sirve como valor_sin_iva si no hay factura
                        if doc.get("ras_valor_dnp") is not None and "valor_sin_iva" not in p:
                            p["valor_sin_iva"] = parse_numero(doc["ras_valor_dnp"])
                        # Usar fecha fin de periodo como fecha del pago si no hay otra
                        if doc.get("ras_periodo_fin") and "fecha" not in p:
                            p["fecha"] = doc["ras_periodo_fin"]
                except (TypeError, ValueError):
                    pass
            if doc.get("ras_periodo_inicio") and not exp["fecha_inicio_ejecucion"]:
                exp["fecha_inicio_ejecucion"] = doc["ras_periodo_inicio"]
            if doc.get("ras_periodo_fin") and not exp["fecha_fin_ejecucion"]:
                exp["fecha_fin_ejecucion"] = doc["ras_periodo_fin"]
            if doc.get("ras_pago_total") and not exp["numero_pagos"]:
                exp["numero_pagos"] = doc["ras_pago_total"]
            if doc.get("ras_nombres_supervisor") and not exp["supervisor"]:
                exp["supervisor"] = doc["ras_nombres_supervisor"]

        # ── FACTURA ────────────────────────────────────────
        elif "FACTURA" in tipo:
            mes = doc.get("mes_cobro_honorarios") or mes_desde_fecha(doc.get("factura_fecha_emision"))
            if mes:
                try:
                    mes_int = int(mes)
                    if 1 <= mes_int <= 12:
                        p = pago_mes(mes_int)
                        if doc.get("factura_valor_sin_iva") is not None: p["valor_sin_iva"]  = parse_numero(doc["factura_valor_sin_iva"])
                        if doc.get("factura_iva")           is not None: p["iva"]            = parse_numero(doc["factura_iva"])
                        if doc.get("factura_total")         is not None: p["total_factura"]  = parse_numero(doc["factura_total"])
                        if doc.get("factura_fecha_emision"):             p["fecha"]          = doc["factura_fecha_emision"]
                        if doc.get("factura_numero"):                    p["numero_factura"] = doc["factura_numero"]
                        if doc.get("factura_cufe"):                      p["cufe"]           = doc["factura_cufe"]
                        # valor_dnp desde factura si no hay RAS
                        if "valor_dnp" not in p and doc.get("factura_valor_sin_iva") is not None:
                            p["valor_dnp"] = parse_numero(doc["factura_valor_sin_iva"])
                except (TypeError, ValueError):
                    pass
            # Si la factura trae IVA > 0, el contratista es responsable de IVA
            iva = parse_numero(doc.get("factura_iva")) or 0
            if iva > 0:
                exp["responsable_iva"] = "Si"

        # ── AFC ────────────────────────────────────────────
        elif "AFC" in tipo:
            mes   = doc.get("afc_mes_aplicable") or mes_desde_fecha(doc.get("afc_fecha_solicitud"))
            monto = parse_numero(doc.get("afc_monto"))
            if mes and monto:
                try:
                    mes_int = int(mes)
                    if 1 <= mes_int <= 12:
                        exp["afc_por_mes"][mes_int] = exp["afc_por_mes"].get(mes_int, 0.0) + monto
                except (TypeError, ValueError):
                    pass
            if doc.get("afc_banco") and not exp["banco"]:
                exp["banco"] = doc["afc_banco"]
            if doc.get("afc_numero_cuenta") and not exp["numero_cuenta"]:
                exp["numero_cuenta"] = doc["afc_numero_cuenta"]

        # ── DEPENDIENTES ───────────────────────────────────
        elif "DEPENDIENTE" in tipo:
            exp["dependientes"].append({
                "nombre"              : doc.get("dep_nombre"),
                "tipo_doc"            : doc.get("dep_tipo_doc"),
                "numero_doc"          : doc.get("dep_numero_doc"),
                "fecha_nacimiento"    : doc.get("dep_fecha_nacimiento"),
                "parentesco"          : doc.get("dep_parentesco"),
                "calidad_dependencia" : doc.get("dep_calidad_dependencia"),
                "fecha_solicitud"     : doc.get("dep_fecha_solicitud"),
            })

        # ── FORMATO 5 ──────────────────────────────────────
        elif "FORMATO" in tipo:
            mes_cobro = doc.get("mes_cobro_honorarios")
            mes_pla   = doc.get("formato5_mes_planilla")
            if mes_cobro and mes_pla:
                try:
                    formato5_index[(cedula, int(mes_cobro))] = int(mes_pla)
                except (TypeError, ValueError):
                    pass
            if mes_cobro and doc.get("formato5_deduccion_renta"):
                try:
                    pago_mes(int(mes_cobro))["deduccion_renta_solicitada"] = doc["formato5_deduccion_renta"]
                except (TypeError, ValueError):
                    pass

    # ─── FASE 2: vincular PILAs huérfanas ──────────────────────────
    for pila in pilas_huerfanas:
        cedula  = str(pila.get("identificacion_contratista") or "SIN_CEDULA").strip()
        exp     = expedientes[cedula]
        exp["identificacion"] = cedula
        if not exp["nombre"] and pila.get("nombre_contratista"):
            exp["nombre"] = pila["nombre_contratista"]
        exp["documentos_detectados"].append({"tipo": "PLANILLA", "archivo": pila.get("_archivo", "")})

        mes_pila = pila.get("pila_periodo_pagado") or mes_desde_fecha(pila.get("pila_fecha_pago"))
        if not mes_pila:
            continue
        mes_pila = int(mes_pila)

        # Prioridad 1: Formato 5 explícito
        mes_cobro_asignado = None
        for (ced, mes_c), mes_p in formato5_index.items():
            if ced == cedula and mes_p == mes_pila:
                mes_cobro_asignado = mes_c
                break

        # Prioridad 2: Pago Vencido (mes_pila + 1 coincide con un mes con pago)
        if not mes_cobro_asignado:
            cand = mes_pila + 1 if mes_pila < 12 else 1
            if cand in exp["pagos_por_mes"]:
                mes_cobro_asignado = cand

        # Prioridad 3: Pago Corriente
        if not mes_cobro_asignado and mes_pila in exp["pagos_por_mes"]:
            mes_cobro_asignado = mes_pila

        exp["pilas_por_mes"][mes_pila] = {
            "fecha_pago"          : pila.get("pila_fecha_pago"),
            "operador"            : pila.get("pila_operador"),
            "ibc"                 : pila.get("pila_ibc"),
            "salud"               : pila.get("pila_salud"),
            "pension"             : pila.get("pila_pension"),
            "arl"                 : pila.get("pila_arl"),
            "mes_cobro_asignado"  : mes_cobro_asignado,
        }

    return dict(expedientes)


# ════════════════════════════════════════════════════════════════════
# AUDITORÍA (REGLAS DE LEY)
# ════════════════════════════════════════════════════════════════════
def auditar_expediente(exp: dict) -> dict:
    """Aplica reglas básicas y devuelve un resumen de estado."""
    alertas = []
    tipos_docs = {d["tipo"] for d in exp.get("documentos_detectados", [])}

    # Completitud mínima
    if not any("RAS" in t for t in tipos_docs):
        alertas.append("❌ Falta RAS")
    if not any("PLANILLA" in t for t in tipos_docs):
        alertas.append("⚠️ Falta PLANILLA")

    # Cruce RAS vs Factura mes a mes
    for mes, pago in exp.get("pagos_por_mes", {}).items():
        ras   = parse_numero(pago.get("valor_dnp")) or 0
        fact  = parse_numero(pago.get("valor_sin_iva")) or 0
        if ras > 0 and fact > 0 and abs(ras - fact) > 10:
            alertas.append(f"❌ Mes {mes}: RAS (${ras:,.0f}) ≠ Factura (${fact:,.0f})")

    # IBC vs SMLMV (Ley)
    for mes, pila in exp.get("pilas_por_mes", {}).items():
        ibc = parse_numero(pila.get("ibc")) or 0
        salud = parse_numero(pila.get("salud")) or 0
        pension = parse_numero(pila.get("pension")) or 0
        if ibc > 0:
            if ibc < (SMLMV_2026 - 1000):
                alertas.append(f"❌ Mes {mes}: IBC bajo (${ibc:,.0f} < SMLMV)")
            if salud > 0 and salud < ((ibc * 0.125) - 1000):
                alertas.append(f"❌ Mes {mes}: Salud < 12.5% IBC")
            if pension > 0 and pension < ((ibc * 0.16) - 1000):
                alertas.append(f"❌ Mes {mes}: Pensión < 16% IBC")

    # Inferencia de riesgo ARL (sobre el último mes con datos)
    riesgo = "N/A"
    pilas = list(exp.get("pilas_por_mes", {}).values())
    if pilas:
        ult = pilas[-1]
        ibc = parse_numero(ult.get("ibc")) or 0
        arl = parse_numero(ult.get("arl")) or 0
        if ibc > 0 and arl > 0:
            pct = (arl / ibc) * 100
            tabla = [(0.522, "I"), (1.044, "II"), (2.436, "III"),
                     (4.350, "IV"), (6.960, "V")]
            for ref, nombre in tabla:
                if abs(pct - ref) < 0.1:
                    riesgo = f"Riesgo {nombre} ({ref}%)"
                    break
            else:
                riesgo = f"Atípico ({pct:.2f}%)"
    exp["clase_riesgo_arl"] = riesgo

    exp["alertas"] = alertas
    exp["estado"]  = "✅ Validado" if not alertas else "⚠️ Con alertas"
    return exp


# ════════════════════════════════════════════════════════════════════
# LLENADO DE LA PLANTILLA EXCEL
# ════════════════════════════════════════════════════════════════════
def llenar_excel(exp: dict, plantilla_bytes: bytes) -> bytes:
    """
    Carga la plantilla, escribe los datos del expediente, devuelve bytes.

    Mapeo fila ← mes: fila = 37 + mes  (Ene=38, Feb=39, …, Dic=49).
    Si un mes no tiene documentos pero es el mes actual de ejecución,
    se escribe igualmente la fecha del último día del mes (para que las
    fórmulas de seguridad social tengan referencia temporal).
    """
    wb = load_workbook(io.BytesIO(plantilla_bytes))
    ws = wb["RT-Cto1"]

    # ── 1) Cabecera (inputs raw en fila 1) ───────────────────────
    if exp.get("nombre"):
        ws["N1"] = exp["nombre"]

    id_str = exp.get("identificacion") or ""
    if exp.get("numero_contrato"):
        id_str = f"{id_str} / Cto. {exp['numero_contrato']}"
    if id_str:
        ws["M1"] = id_str

    if exp.get("dependencia"):
        ws["U1"] = exp["dependencia"]
    if exp.get("supervisor"):
        ws["AW1"] = exp["supervisor"]
    if exp.get("banco"):
        ws["AH1"] = exp["banco"]
    if exp.get("numero_cuenta"):
        ws["P1"] = exp["numero_cuenta"]

    # Fechas y valores del contrato
    f_ini = parse_fecha(exp.get("fecha_inicio_ejecucion"))
    f_fin = parse_fecha(exp.get("fecha_fin_ejecucion"))
    if f_ini: ws["AT1"] = f_ini
    if f_fin: ws["AU1"] = f_fin

    valor_mensual_contrato = None
    if exp.get("valor_mensual"):
        valor_mensual_contrato = parse_numero(exp["valor_mensual"])
        if valor_mensual_contrato is not None:
            ws["BB1"] = valor_mensual_contrato
    if exp.get("valor_total"):
        v = parse_numero(exp["valor_total"])
        if v is not None: ws["AY1"] = v
    if exp.get("numero_pagos"):
        try: ws["AV1"] = int(exp["numero_pagos"])
        except (TypeError, ValueError): pass

    # ── 2) Configuración tributaria ───────────────────────────────
    ws["H12"] = "Si" if exp.get("es_pensionado")   == "Si" else "No"
    ws["H13"] = "Si" if exp.get("responsable_iva") == "Si" else "No"
    ws["H14"] = "Si" if exp.get("regimen_simple")  == "Si" else "No"

    # ── 3) Grilla mensual ─────────────────────────────────────────
    # Mes actual: si corro el script hoy, el mes a reportar es el
    # mes ANTERIOR (el pago ya ocurrió). Si estamos en junio → mayo.
    # Pero si hay documentos del mes actual, los usamos tal cual.
    hoy = datetime.now()
    mes_ejecucion = hoy.month           # mes del que se reporta (corriente)
    anio_ejecucion = hoy.year

    pagos = exp.get("pagos_por_mes") or {}

    # Si no hay ningún pago con fecha/valor, insertar fila para el mes actual
    # SOLO si hay documentos que lo justifican (dependientes o AFC del mes)
    tiene_docs_utiles = (
        bool(exp.get("dependientes"))
        or bool(exp.get("afc_por_mes"))
        or bool(exp.get("pilas_por_mes"))
    )
    if not pagos and tiene_docs_utiles:
        import calendar
        ult_dia = calendar.monthrange(anio_ejecucion, mes_ejecucion)[1]
        pagos = {mes_ejecucion: {
            "fecha": datetime(anio_ejecucion, mes_ejecucion, ult_dia),
        }}

    for mes, pago in pagos.items():
        try:
            mes_int = int(mes)
            if not 1 <= mes_int <= 12:
                continue
        except (TypeError, ValueError):
            continue

        # ↓ MAPEO CORRECTO: fila = 37 + mes (1=38, 2=39, …, 12=49)
        fila = 37 + mes_int

        # Fecha del pago (R): viene de factura, RAS o se construye desde el mes
        fecha = (parse_fecha(pago.get("fecha"))
                 or parse_fecha(pago.get("fecha_registro")))
        if fecha is None and isinstance(pago.get("fecha"), datetime):
            fecha = pago["fecha"]
        if fecha:
            ws[f"R{fila}"] = fecha
        else:
            # fallback: último día del mes del documento
            import calendar
            ult_dia = calendar.monthrange(anio_ejecucion, mes_int)[1]
            ws[f"R{fila}"] = datetime(anio_ejecucion, mes_int, ult_dia)

        # Valor sin IVA (T): columna clave — necesaria para calcular dependientes
        valor = (parse_numero(pago.get("valor_sin_iva"))
                 or parse_numero(pago.get("valor_dnp")))
        if valor is not None:
            ws[f"T{fila}"] = valor

        # Registro contable (Q)
        if pago.get("registro_contable_no"):
            ws[f"Q{fila}"] = pago["registro_contable_no"]

    # AFC por mes
    for mes_k, monto in (exp.get("afc_por_mes") or {}).items():
        try:
            mes_int = int(mes_k)
            if not 1 <= mes_int <= 12:
                continue
        except (TypeError, ValueError):
            continue
        monto_f = parse_numero(monto)
        if monto_f is None or monto_f <= 0:
            continue
        ws[f"AC{37 + mes_int}"] = monto_f

    # Si hay dependientes declarados pero no hay pago registrado para el mes
    # actual, insertar fila con fecha del último día del mes. Esto activa
    # la fórmula AA{fila} → "SI" para que el usuario vea la deducción en cuanto
    # agregue el valor T manualmente (o en la próxima corrida con factura).
    import calendar as _cal
    deps_activos = bool(exp.get("dependientes"))
    if deps_activos:
        mes_actual_int = hoy.month
        if mes_actual_int not in pagos:
            ult_dia = _cal.monthrange(anio_ejecucion, mes_actual_int)[1]
            fila_actual = 37 + mes_actual_int
            if ws[f"R{fila_actual}"].value is None:
                ws[f"R{fila_actual}"] = datetime(anio_ejecucion, mes_actual_int, ult_dia)

    # ── 4) H18: valor mensual sin IVA ────────────────────────────
    # H18 es el DENOMINADOR de la fórmula de Dependientes Económicos.
    # La plantilla lo calcula con =IF(H13="SI", ROUND(H17/1.19,0), H17).
    # H17 apunta a BB1. Si no tenemos BB1, intentamos inferirlo
    # desde el primer valor de pago que tengamos.
    if valor_mensual_contrato is None and pagos:
        valores = [parse_numero(p.get("valor_sin_iva")) or parse_numero(p.get("valor_dnp"))
                   for p in pagos.values()]
        valores = [v for v in valores if v]
        if valores:
            # No escribimos H18 directamente (tiene fórmula);
            # escribimos BB1 para que H17 → H18 → AB se calculen solos.
            ws["BB1"] = max(valores)

    # ── 5) Panel de dependientes ──────────────────────────────────
    deps = exp.get("dependientes") or []
    if deps:
        dep = deps[0]
        if dep.get("tipo_doc"):
            ws["AH10"] = dep["tipo_doc"]
        if dep.get("numero_doc"):
            ws["AI10"] = dep["numero_doc"]

        fn = parse_fecha(dep.get("fecha_nacimiento"))
        fs = parse_fecha(dep.get("fecha_solicitud"))
        if fn: ws["AI11"] = fn
        if fs: ws["AI12"] = fs

        # AG6 / AG9 activan la fórmula AA{fila} → "SI" / "No"
        # que a su vez activa el cálculo de AB (dependientes)
        if fs:
            ws["AG6"] = fs

        calidad = (dep.get("calidad_dependencia") or "").lower()
        fecha_hasta = None
        if "menor" in calidad and fn:
            fecha_hasta = add_years_safe(fn, 18)
        elif "estudiante" in calidad and fn:
            fecha_hasta = add_years_safe(fn, 25)
        elif fs:
            fecha_hasta = add_years_safe(fs, 1)
        if fecha_hasta:
            ws["AG9"] = fecha_hasta

    # Forzar recálculo al abrir
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def empaquetar_zip(excels: dict) -> bytes:
    """excels = {nombre_archivo: bytes}. Devuelve bytes de un ZIP."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, contenido in excels.items():
            zf.writestr(nombre, contenido)
    buffer.seek(0)
    return buffer.getvalue()


# ════════════════════════════════════════════════════════════════════
# UI — KPIs, GRÁFICAS, PROGRESO
# ════════════════════════════════════════════════════════════════════
PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="IBM Plex Mono, monospace", color="#6b8cc4", size=11),
    margin=dict(l=10, r=10, t=40, b=10),
)
COLOR_PALETTE = ["#3b82f6", "#1d4ed8", "#60a5fa", "#93c5fd", "#2563eb",
                 "#1e40af", "#bfdbfe", "#0ea5e9", "#38bdf8", "#7dd3fc"]


def render_progress(current: int, total: int, filename: str, mensaje: str, slot):
    pct = int((current / total) * 100) if total > 0 else 0
    slot.markdown(f"""
    <div class="progress-wrapper">
        <div class="progress-title">⬡ Procesando archivos</div>
        <div class="progress-track">
            <div class="progress-fill" style="width:{pct}%"></div>
        </div>
        <div class="progress-pulse">
            <div class="pulse-dot"></div>
            <div class="pulse-dot"></div>
            <div class="pulse-dot"></div>
            <span class="progress-msg">{mensaje}</span>
        </div>
        <div class="progress-counter">{current} / {total} archivos &nbsp;·&nbsp; {pct}% completado</div>
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.65rem;color:#1e3a6e;
                    margin-top:0.4rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">
            ▶ {filename}
        </div>
    </div>
    """, unsafe_allow_html=True)


def kpi_card(value: str, label: str, sub: str = "") -> str:
    sub_html = f"<div class='kpi-sub'>{sub}</div>" if sub else ""
    return f"""
    <div class="kpi-card">
        <div class="kpi-value">{value}</div>
        <div class="kpi-label">{label}</div>
        {sub_html}
    </div>
    """


MENSAJES_PROGRESO = [
    "Enviando a Azure Document Intelligence…",
    "Extrayendo tablas y estructuras…",
    "Analizando imagen con GPT-4o…",
    "Clasificando tipo de documento…",
    "Identificando contratista y montos…",
    "Aplicando vinculación relacional…",
    "Cruzando RAS con facturas…",
    "Verificando IBC vs SMLMV…",
    "Llenando plantilla del contratista…",
    "Empaquetando expedientes…",
]


def grafica_dona_estado(df: pd.DataFrame):
    if df.empty or "Estado" not in df.columns:
        return go.Figure()
    conteo = df["Estado"].apply(
        lambda x: "✅ Validados" if "✅" in str(x) else "⚠️ Con alertas"
    ).value_counts().reset_index()
    conteo.columns = ["Estado", "Cantidad"]
    fig = go.Figure(go.Pie(
        labels=conteo["Estado"], values=conteo["Cantidad"], hole=0.62,
        marker=dict(colors=["#3b82f6", "#1e3a6e"],
                    line=dict(color="#070c18", width=2)),
        textfont=dict(family="IBM Plex Mono, monospace", size=11),
    ))
    fig.update_layout(
        **PLOTLY_LAYOUT,
        title=dict(text="Estado de Expedientes", font=dict(color="#3b82f6", size=13)),
        legend=dict(font=dict(color="#6b8cc4")),
        annotations=[dict(text=f"{conteo['Cantidad'].sum()}<br>total",
                          x=0.5, y=0.5, font_size=13,
                          font=dict(family="IBM Plex Mono", color="#3b82f6"),
                          showarrow=False)],
    )
    return fig


def grafica_montos_contratista(df: pd.DataFrame):
    if df.empty or "Total Pagado" not in df.columns:
        return go.Figure()
    d = df.copy()
    d["Total_M"] = (pd.to_numeric(d["Total Pagado"], errors="coerce").fillna(0) / 1_000_000).round(2)
    d = d.sort_values("Total_M", ascending=True).tail(15)
    fig = px.bar(d, x="Total_M", y="Contratista", orientation="h",
                 color="Total_M",
                 color_continuous_scale=["#1e3a6e", "#3b82f6", "#93c5fd"],
                 title="Top contratistas por monto (millones COP)",
                 labels={"Total_M": "Millones COP"})
    fig.update_traces(marker_line_width=0)
    fig.update_layout(**PLOTLY_LAYOUT, coloraxis_showscale=False, showlegend=False,
                      title_font=dict(size=13, color="#3b82f6"),
                      xaxis=dict(gridcolor="#0d1525", linecolor="#1a2744"),
                      yaxis=dict(gridcolor="rgba(0,0,0,0)", linecolor="#1a2744"))
    return fig


def grafica_documentos_tipo(documentos: list):
    if not documentos:
        return go.Figure()
    df = pd.DataFrame([{"tipo": d.get("_tipo") or "OTRO"} for d in documentos])
    conteo = df["tipo"].value_counts().reset_index()
    conteo.columns = ["Tipo", "Cantidad"]
    fig = px.bar(conteo, x="Tipo", y="Cantidad", color="Tipo",
                 color_discrete_sequence=COLOR_PALETTE,
                 title="Documentos clasificados por tipo")
    fig.update_traces(marker_line_width=0, opacity=0.9)
    fig.update_layout(**PLOTLY_LAYOUT, showlegend=False,
                      title_font=dict(size=13, color="#3b82f6"),
                      xaxis=dict(gridcolor="#0d1525", linecolor="#1a2744"),
                      yaxis=dict(gridcolor="#0d1525", linecolor="#1a2744"))
    return fig


# ════════════════════════════════════════════════════════════════════
# APP STREAMLIT — MAIN
# ════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="Herramienta OCR-SF-UCD V13",
        page_icon="⬡",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    st.markdown("""
    <div class="ocr-header">
        <div class="ocr-badge">DNP · SF · UCD · V13.0</div>
        <h1>⬡ Herramienta OCR-SF-UCD</h1>
        <div class="subtitle">Procesamiento integral · Vinculación relacional · Llenado automático de plantilla RT</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Validación de credenciales ───────────────────────────────
    if not all([ENDPOINT, API_KEY, AZURE_OPENAI_ENDPOINT,
                AZURE_OPENAI_API_KEY, AZURE_OPENAI_DEPLOYMENT]):
        st.error(
            "⚠ Variables de entorno incompletas. Configura "
            "AZURE_FORM_RECOGNIZER_ENDPOINT, AZURE_FORM_RECOGNIZER_KEY, "
            "AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY y "
            "AZURE_OPENAI_DEPLOYMENT_NAME."
        )
        st.stop()

    # ── Validación de plantilla ──────────────────────────────────
    plantilla_bytes = cargar_plantilla_bytes()
    if not plantilla_bytes:
        st.warning(
            f"⚠ No se encontró la plantilla en `{RUTA_PLANTILLA}`. "
            "Súbela manualmente para esta sesión:"
        )
        plantilla_subida = st.file_uploader(
            "Plantilla RT-Cto1.xlsx",
            type=["xlsx"],
            key="plantilla_upload",
        )
        if not plantilla_subida:
            st.stop()
        plantilla_bytes = plantilla_subida.read()

    cliente_di, cliente_oai = get_clients()

    # ── Carga del ZIP ────────────────────────────────────────────
    st.markdown('<div class="section-title">① Carga del ZIP de expedientes</div>',
                unsafe_allow_html=True)
    uploaded_zip = st.file_uploader(
        "Arrastra o selecciona un ZIP con todos los PDFs (RAS, PILA, Facturas, AFC, Dependientes, Formato 5).",
        type=["zip"],
    )

    if not uploaded_zip:
        st.markdown("""
        <div style="text-align:center;padding:3rem 0;color:#1e3a6e;
                    font-family:'IBM Plex Mono',monospace;font-size:0.8rem;">
            Esperando archivo ZIP…<br>
            <span style="font-size:0.65rem;letter-spacing:0.1em;color:#111e38;">
            [ Mezcla libremente PDFs de varios contratistas — se agrupan automáticamente por cédula ]
            </span>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Extracción del ZIP ───────────────────────────────────────
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(uploaded_zip.read()), "r") as zf:
            zf.extractall(tmpdir)

        archivos_pdf = []
        for root, _, files in os.walk(tmpdir):
            for f in files:
                if f.lower().endswith(".pdf"):
                    archivos_pdf.append(os.path.join(root, f))

        total_pdf = len(archivos_pdf)
        if total_pdf == 0:
            st.warning("El ZIP no contiene archivos PDF.")
            return

        st.markdown(f"""
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.75rem;
                    color:#3b82f6;padding:0.6rem 1rem;background:#0a1022;
                    border:1px solid #1a2744;border-radius:6px;margin-bottom:1rem;">
            ✓ ZIP detectado &nbsp;·&nbsp; <strong>{total_pdf}</strong> archivos PDF encontrados
        </div>
        """, unsafe_allow_html=True)

        if not st.button("▶ Iniciar análisis", use_container_width=False):
            return

        # ── FASE A: EXTRACCIÓN DOCUMENTO POR DOCUMENTO ───────────
        st.markdown('<div class="section-title">② Procesamiento</div>',
                    unsafe_allow_html=True)
        progress_slot = st.empty()
        documentos_procesados = []

        for idx, ruta in enumerate(archivos_pdf):
            nombre = os.path.basename(ruta)
            msg_idx = idx % len(MENSAJES_PROGRESO)
            render_progress(idx, total_pdf, nombre,
                            MENSAJES_PROGRESO[msg_idx], progress_slot)

            try:
                texto = extraer_texto_con_azure(ruta, cliente_di)
                render_progress(idx, total_pdf, nombre,
                                "Analizando con GPT-4o…", progress_slot)
                img_b64 = pdf_primera_pagina_a_base64(ruta)
                datos = procesar_con_ia(texto, img_b64, cliente_oai)
                datos["_archivo"] = nombre
                datos["_tipo"]    = str(datos.get("tipo_documento_detectado") or "OTRO").upper()
                documentos_procesados.append(datos)
            except Exception as e:
                documentos_procesados.append({
                    "_archivo"     : nombre,
                    "_tipo"        : "ERROR",
                    "_error_json"  : f"Error técnico: {str(e)[:150]}",
                })

            time.sleep(0.2)

        render_progress(total_pdf, total_pdf, "Vinculando expedientes",
                        "✓ Documentos clasificados", progress_slot)
        time.sleep(0.5)

        # ── FASE B: VINCULACIÓN ──────────────────────────────────
        expedientes = construir_expedientes(documentos_procesados)
        for cedula, exp in expedientes.items():
            auditar_expediente(exp)

        # ── FASE C: LLENADO DE EXCELS ────────────────────────────
        render_progress(total_pdf, total_pdf, "Generando Excels",
                        "Llenando plantillas…", progress_slot)
        excels = {}
        for cedula, exp in expedientes.items():
            try:
                contenido = llenar_excel(exp, plantilla_bytes)
                excels[f"RT-Cto-{cedula}.xlsx"] = contenido
            except Exception as e:
                exp["alertas"].append(f"❌ Error generando Excel: {str(e)[:120]}")

        progress_slot.empty()

        # ── RESUMEN PARA TABLA ───────────────────────────────────
        filas_resumen = []
        for cedula, exp in expedientes.items():
            tipos = {d["tipo"] for d in exp.get("documentos_detectados", [])}
            total_pagado = sum(
                parse_numero(p.get("valor_sin_iva")) or parse_numero(p.get("valor_dnp")) or 0
                for p in exp.get("pagos_por_mes", {}).values()
            )
            total_afc = sum((exp.get("afc_por_mes") or {}).values())
            n_dep = len(exp.get("dependientes") or [])
            n_meses = len(exp.get("pagos_por_mes") or {})
            filas_resumen.append({
                "Contratista"      : exp.get("nombre") or "—",
                "Identificación"   : cedula,
                "Contrato"         : exp.get("numero_contrato") or "—",
                "Meses con pago"   : n_meses,
                "Total Pagado"     : round(total_pagado, 0),
                "Total AFC"        : round(total_afc, 0),
                "Dependientes"     : n_dep,
                "Documentos"       : ", ".join(sorted(tipos)),
                "Riesgo ARL"       : exp.get("clase_riesgo_arl", "N/A"),
                "Estado"           : exp.get("estado", "—"),
                "Alertas"          : " | ".join(exp.get("alertas", [])) or "—",
            })

        df = pd.DataFrame(filas_resumen)

        # ── BANNER ────────────────────────────────────────────────
        total_contratistas = len(df)
        total_dinero = df["Total Pagado"].sum() if not df.empty else 0
        validados = (df["Estado"].str.contains("✅").sum() if not df.empty else 0)

        st.markdown(f"""
        <div class="success-banner">
            ✓ Análisis completado &nbsp;·&nbsp;
            {total_pdf} PDFs procesados &nbsp;·&nbsp;
            {total_contratistas} expedientes generados &nbsp;·&nbsp;
            {validados} validados &nbsp;·&nbsp;
            Total: ${total_dinero/1_000_000:,.1f}M COP
        </div>
        """, unsafe_allow_html=True)

        # ── KPIs ──────────────────────────────────────────────────
        st.markdown('<div class="section-title">③ Indicadores clave</div>',
                    unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(kpi_card(str(total_pdf), "PDFs procesados",
                             sub="Documentos leídos"), unsafe_allow_html=True)
        c2.markdown(kpi_card(str(total_contratistas), "Expedientes",
                             sub="Excels generados"), unsafe_allow_html=True)
        c3.markdown(kpi_card(f"${total_dinero/1_000_000:,.1f}M",
                             "Total pagado", sub="Suma sin IVA · COP"),
                    unsafe_allow_html=True)
        c4.markdown(kpi_card(str(validados),
                             "Validados sin alertas",
                             sub=f"de {total_contratistas}"),
                    unsafe_allow_html=True)

        # ── GRÁFICAS ──────────────────────────────────────────────
        st.markdown('<div class="section-title">④ Análisis visual</div>',
                    unsafe_allow_html=True)
        cg1, cg2 = st.columns([1.4, 1])
        with cg1:
            st.plotly_chart(grafica_documentos_tipo(documentos_procesados),
                            use_container_width=True)
        with cg2:
            st.plotly_chart(grafica_dona_estado(df), use_container_width=True)
        st.plotly_chart(grafica_montos_contratista(df), use_container_width=True)

        # ── TABLA ─────────────────────────────────────────────────
        st.markdown('<div class="section-title">⑤ Resumen de expedientes</div>',
                    unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True, hide_index=True)

        # ── DESCARGA ZIP ──────────────────────────────────────────
        st.markdown('<div class="section-title">⑥ Exportar</div>',
                    unsafe_allow_html=True)

        if excels:
            zip_bytes = empaquetar_zip(excels)
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            st.download_button(
                label=f"⬇ Descargar ZIP con {len(excels)} Excels (RT-Cto-*.xlsx)",
                data=zip_bytes,
                file_name=f"Expedientes_RT_DNP_{ts}.zip",
                mime="application/zip",
            )
        else:
            st.error("No se pudo generar ningún Excel.")

        st.markdown("""
        <div style="font-family:'IBM Plex Mono',monospace;font-size:0.62rem;
                    color:#1a2744;text-align:center;margin-top:3rem;padding-top:1rem;
                    border-top:1px solid #0d1525;">
            Herramienta OCR-SF-UCD · V13.0 · DNP Colombia · Subdirección Financiera
        </div>
        """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()